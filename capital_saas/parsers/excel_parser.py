import re
from pathlib import Path

from openpyxl import load_workbook

from parsers.base_parser import BaseParser


FIELD_ALIASES = {
    "annual_revenue": ["营业收入", "主营业务收入", "销售收入", "营收"],
    "net_profit": ["净利润", "税后利润"],
    "total_assets": ["资产总计", "总资产"],
    "debt_total": ["负债合计", "总负债"],
    "monthly_cashflow": ["经营活动现金流量净额", "经营现金流", "现金流"],
    "accounts_receivable": ["应收账款", "应收款"],
    "receivable_days": ["应收账款周期", "应收天数", "账期天数"],
    "short_debt": ["短期借款", "短期负债"],
    "long_debt": ["长期借款", "长期负债"],
    "tax_amount": ["纳税金额", "应交税费", "实缴税额"],
}


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).replace(",", "").replace("，", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group())
    if "亿" in text:
        number *= 100_000_000
    elif "万" in text:
        number *= 10_000
    return number


class ExcelParser(BaseParser):
    parser_type = "excel"

    def parse(self, file_path: Path) -> dict:
        if file_path.suffix.lower() == ".xls":
            raise ValueError("旧版 .xls 暂不支持，请另存为 .xlsx 后重新上传")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        previews = {}
        fields = {}
        field_sources = {}
        for sheet in workbook.worksheets:
            rows = []
            for values in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
                row = [value for value in values[:15]]
                if any(value not in (None, "") for value in row):
                    rows.append(row)
                for index, value in enumerate(row):
                    label = str(value or "").strip()
                    for key, aliases in FIELD_ALIASES.items():
                        if key in fields or not any(alias in label for alias in aliases):
                            continue
                        candidates = row[index + 1:] + ([] if not rows else rows[-1][index + 1:])
                        for candidate in candidates:
                            number = _number(candidate)
                            if number is not None:
                                fields[key] = number
                                field_sources[key] = {"sheet": sheet.title, "label": label}
                                break
            previews[sheet.title] = rows[:20]
        workbook.close()
        return {
            "parser_type": self.parser_type,
            "status": "success",
            "sheet_names": list(previews),
            "preview_rows": previews,
            "financial_fields": fields,
            "field_sources": field_sources,
        }
