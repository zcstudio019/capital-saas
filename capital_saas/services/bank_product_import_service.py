"""CSV, Excel and Markdown bank-product import service."""
from __future__ import annotations
import csv
import io
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from db.models import BankProduct
from parsers.markdown_bank_product_parser import normalize_bank_product, parse_markdown_bank_products

class ImportResult(dict):
    def __eq__(self, other):
        if isinstance(other, dict):
            return all(self.get(key) == value for key, value in other.items())
        return super().__eq__(other)

def parse_bank_product_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".md":
        return parse_markdown_bank_products(content)
    if suffix == ".csv":
        return [normalize_bank_product(row) for row in csv.DictReader(io.StringIO(content.decode("utf-8-sig")))]
    if suffix == ".xlsx":
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
        values = sheet.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in next(values, [])]
        return [normalize_bank_product(dict(zip(headers, row))) for row in values if any(v is not None for v in row)]
    raise ValueError("仅支持 CSV、Excel（.xlsx）或 Markdown（.md）文件")

def import_bank_products(db: Session, rows: list[dict[str, Any]], source_file_name: str = "") -> dict[str, Any]:
    batch_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:8]
    result: dict[str, Any] = ImportResult({"parsed": len(rows), "success": 0, "created": 0, "updated": 0, "failed": 0, "errors": [], "batch_id": batch_id, "products": []})
    seen: set[str] = set()
    model_fields = set(BankProduct.__table__.columns.keys())
    for index, row in enumerate(rows, 1):
        identifier = row.get("product_code") or row.get("_title") or row.get("product_name") or f"第{index}项"
        if not row.get("product_name") or not row.get("bank_name"):
            result["failed"] += 1
            result["errors"].append(f"{identifier}：缺少银行/机构或产品名称")
            continue
        dedupe = row.get("product_code") or f'{row["bank_name"]}|{row["product_name"]}'
        if dedupe in seen:
            result["failed"] += 1
            result["errors"].append(f"{identifier}：文件内产品重复")
            continue
        seen.add(dedupe)
        item = None
        if row.get("product_code"):
            item = db.query(BankProduct).filter(BankProduct.product_code == row["product_code"]).first()
        if item is None:
            item = db.query(BankProduct).filter(BankProduct.bank_name == row["bank_name"], BankProduct.product_name == row["product_name"]).first()
        created = item is None
        if created:
            item = BankProduct()
            db.add(item)
        try:
            for field, value in row.items():
                if field in model_fields and field not in {"id", "created_at", "updated_at"}:
                    setattr(item, field, value)
            # Database stores yuan; parser exposes the requested 万元 normalization.
            item.min_amount = (row.get("min_amount") or 0) * 10000
            item.max_amount = (row.get("max_amount") or 0) * 10000
            item.bank_type = row.get("bank_type") or row.get("institution_category") or "银行"
            item.product_type = row.get("product_type") or "待补充"
            item.suitable_industry = row.get("suitable_industry") or "通用"
            item.application_requirements = row.get("access_conditions_json", "")
            item.required_documents = row.get("required_documents_json", "")
            item.data_source, item.is_active = "imported", True
            item.source_file_name, item.source_batch_id, item.imported_at = source_file_name, batch_id, datetime.now()
            result["success"] += 1
            result["created" if created else "updated"] += 1
            result["products"].append({"product_code": row.get("product_code", ""), "product_name": row["product_name"]})
        except Exception as exc:
            result["failed"] += 1
            result["errors"].append(f"{identifier}：{exc}")
    if result["success"]:
        db.commit()
    else:
        db.rollback()
    return result

def disable_mock_products(db: Session) -> int:
    count = db.query(BankProduct).filter(BankProduct.data_source == "mock", BankProduct.is_active.is_(True)).update({BankProduct.is_active: False}, synchronize_session=False)
    db.commit()
    return count
